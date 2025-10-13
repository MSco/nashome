from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
import xml.etree.ElementTree as ET

def hex_to_rgb_tuple(hexstr: str):
    hexstr = hexstr.strip("#")
    if len(hexstr) == 6:
        return tuple(int(hexstr[i:i+2], 16) for i in (0, 2, 4))
    raise ValueError("Ungültiges Hex-Format")

def rgb_tuple_to_hex(rgb):
    return "#{:02X}{:02X}{:02X}".format(
        max(0, min(255, int(round(rgb[0])))),
        max(0, min(255, int(round(rgb[1])))),
        max(0, min(255, int(round(rgb[2]))))
    )

def parse_theme_colors_from_workbook(wb):
    theme_map = {}
    loaded_theme = getattr(wb, "loaded_theme", None)
    if not loaded_theme:
        return theme_map
    try:
        root = ET.fromstring(loaded_theme)
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        clr_scheme = root.find(".//a:clrScheme", ns)
        if clr_scheme is None:
            return theme_map
        for child in clr_scheme:
            tag_name = child.tag.split("}")[-1]
            srgb = child.find(".//a:srgbClr", ns)
            if srgb is not None and "val" in srgb.attrib:
                theme_map[tag_name] = f"#{srgb.attrib['val']}"
            else:
                sysclr = child.find(".//a:sysClr", ns)
                if sysclr is not None and "lastClr" in sysclr.attrib:
                    theme_map[tag_name] = f"#{sysclr.attrib['lastClr']}"
        return theme_map
    except Exception:
        return theme_map

def excel_color_to_hex(wb, color_obj):
    fallback = "#FFFFFF"
    if color_obj is None:
        return fallback
    rgb_attr = getattr(color_obj, "rgb", None)
    if isinstance(rgb_attr, str):
        rgb_attr = rgb_attr.strip()
        if len(rgb_attr) == 8:  # AARRGGBB
            rgb_attr = rgb_attr[2:]
        if not rgb_attr.startswith("#"):
            rgb_attr = "#" + rgb_attr
        return rgb_attr.upper()
    theme = getattr(color_obj, "theme", None)
    tint = getattr(color_obj, "tint", None)
    if theme is not None:
        if not hasattr(wb, "_theme_color_map_cache"):
            theme_colors = parse_theme_colors_from_workbook(wb)
            ordered_keys = ["lt1","dk1","lt2","dk2","accent1","accent2","accent3","accent4","accent5","accent6"]
            wb._theme_color_map_cache = [theme_colors.get(k) for k in ordered_keys]
        theme_list = getattr(wb, "_theme_color_map_cache", [])
        try:
            idx = int(theme)
            if 0 <= idx < len(theme_list):
                base = theme_list[idx]
                if base:
                    if tint:
                        return base  # Tint wird hier nicht angewendet
                    return base
        except Exception:
            pass
    idx = getattr(color_obj, "indexed", None)
    if idx is not None:
        try:
            val = COLOR_INDEX[int(idx)]
            if isinstance(val, str):
                if len(val) == 8:
                    val = val[2:]
                return f"#{val.upper()}"
        except Exception:
            pass
    return fallback

def get_cell_color(wb, cell):
    fill = cell.fill
    if not fill:
        return "#FFFFFF"
    color = None
    if fill.patternType == "solid":
        color = excel_color_to_hex(wb, fill.fgColor)
    else:
        color = excel_color_to_hex(wb, fill.bgColor)
    if (
        color in (None, "", "#000000", "#00000000", "#FF000000")
        or "None" in str(color)
        or color.upper().startswith("#00") and color.upper() not in ("#00FFFF", "#00FF00", "#0000FF")
    ):
        color = "#FFFFFF"
    return color

def read_excel_with_colors(filename, sheet_name=None, start_row=2, end_row=28):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    data, colors = [], []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        row_data, row_colors = [], []
        for cell in row:
            row_data.append(cell.value)
            row_colors.append(get_cell_color(wb, cell))
        data.append(row_data)
        colors.append(row_colors)
    return data, colors

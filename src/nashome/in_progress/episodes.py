from flask import Flask, render_template_string
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
import xml.etree.ElementTree as ET
import re

app = Flask(__name__)

EXCEL_FILE = "data/episodes/season_numbering.xlsx"

# --- Blockdefinition für alle Staffeln ---
BLOCKS = [
    ("EP", 1, 82, 1),
    ("EP", 83, 118, 2),
    ("EP", 119, 159, 3),
    ("EP", 160, 211, 4),
    ("EP", 212, 276, 5),
    ("AG", 1, 40, 6),
    ("AG", 41, 92, 7),
    ("AG", 93, 145, 8),
    ("AG", 146, 192, 9),
    ("DP", 1, 52, 10),
    ("DP", 53, 104, 11),
    ("DP", 105, 157, 12),
    ("DP", 158, 191, 13),
    ("BW", 1, 48, 14),
    ("BW", 49, 97, 15),
    ("BW", 98, 142, 16),
    ("XY", 1, 49, 17),
    ("XY", 50, 93, 18),
    ("XY", 94, 140, 19),
    ("SM", 1, 43, 20),
    ("SM", 44, 92, 21),
    ("SM", 93, 146, 22),
    ("PM", 1, 48, 23),
    ("PM", 49, 90, 24),
    ("PM", 91, 147, 25),
]

# --- Funktionen für Episodenumwandlung ---
def parse_episode_to_season_ep(code):
    """Wandelt EP###/AG### in (season, ep_in_season)"""
    if not code:
        return None, None
    code = str(code).upper().strip()
    m = re.match(r"([A-Z]+)(\d+)", code)
    if not m:
        return None, None
    prefix, num_str = m.groups()
    num = int(num_str)
    for pfx, start, end, season in BLOCKS:
        if prefix == pfx and start <= num <= end:
            return season, num - start + 1
    return None, None

def revert_episode_code(season: int, new_ep_num: int) -> str:
    """Wandelt season + ep_in_season zurück in EP###/AG###"""
    for pfx, start, end, blk_season in BLOCKS:
        if blk_season == season:
            ep_num = start + new_ep_num - 1
            if ep_num > end:
                raise ValueError(f"Episodennummer {new_ep_num} überschreitet das Ende der Staffel {season}")
            return f"{pfx}{ep_num:03d}"
    raise ValueError(f"Keine passende Staffel für {season} gefunden.")

# --- Excel-Farbverarbeitung ---
def hex_to_rgb_tuple(hexstr: str):
    hexstr = hexstr.strip("#")
    if len(hexstr) == 6:
        return tuple(int(hexstr[i:i+2], 16) for i in (0, 2, 4))
    raise ValueError("Ungültiges Hex-Format")

def rgb_tuple_to_hex(rgb):
    return "#{:02X}{:02X}{:02X}".format(
        max(0, min(255, int(round(rgb[0])))),
        max(0, min(255, int(round(rgb[1])))),
        max(0, min(255, int(round(rgb[2]))))
    )

def parse_theme_colors_from_workbook(wb):
    theme_map = {}
    loaded_theme = getattr(wb, "loaded_theme", None)
    if not loaded_theme:
        return theme_map
    try:
        root = ET.fromstring(loaded_theme)
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        clr_scheme = root.find(".//a:clrScheme", ns)
        if clr_scheme is None:
            return theme_map
        for child in clr_scheme:
            tag_name = child.tag.split("}")[-1]
            srgb = child.find(".//a:srgbClr", ns)
            if srgb is not None and "val" in srgb.attrib:
                theme_map[tag_name] = f"#{srgb.attrib['val']}"
            else:
                sysclr = child.find(".//a:sysClr", ns)
                if sysclr is not None and "lastClr" in sysclr.attrib:
                    theme_map[tag_name] = f"#{sysclr.attrib['lastClr']}"
        return theme_map
    except Exception:
        return theme_map


def excel_color_to_hex(wb, color_obj):
    fallback = "#FFFFFF"
    if color_obj is None:
        return fallback

    rgb_attr = getattr(color_obj, "rgb", None)
    if isinstance(rgb_attr, str):
        rgb_attr = rgb_attr.strip()
        if len(rgb_attr) == 8:  # AARRGGBB
            rgb_attr = rgb_attr[2:]
        if not rgb_attr.startswith("#"):
            rgb_attr = "#" + rgb_attr
        return rgb_attr.upper()

    # Theme-Farben
    theme = getattr(color_obj, "theme", None)
    tint = getattr(color_obj, "tint", None)
    if theme is not None:
        if not hasattr(wb, "_theme_color_map_cache"):
            theme_colors = parse_theme_colors_from_workbook(wb)
            ordered_keys = ["lt1","dk1","lt2","dk2","accent1","accent2","accent3","accent4","accent5","accent6"]
            wb._theme_color_map_cache = [theme_colors.get(k) for k in ordered_keys]
        theme_list = getattr(wb, "_theme_color_map_cache", [])
        try:
            idx = int(theme)
            if 0 <= idx < len(theme_list):
                base = theme_list[idx]
                if base:
                    if tint:
                        return apply_tint(base, tint)
                    return base
        except Exception:
            pass

    # Indexed color (wie bisher)
    idx = getattr(color_obj, "indexed", None)
    if idx is not None:
        try:
            val = COLOR_INDEX[int(idx)]
            if isinstance(val, str):
                if len(val) == 8:
                    val = val[2:]
                return f"#{val.upper()}"
        except Exception:
            pass

    return fallback



def get_cell_color(wb, cell):
    """Bestimmt die tatsächlich sichtbare Farbe einer Zelle."""
    fill = cell.fill
    if not fill:
        return "#FFFFFF"

    color = None

    # Nur "solid" pattern wirklich einfärben, Rest = weiß
    if fill.patternType == "solid":
        color = excel_color_to_hex(wb, fill.fgColor)
    else:
        color = excel_color_to_hex(wb, fill.bgColor)

    # Ungültige oder transparente Farben als weiß behandeln
    if (
        color in (None, "", "#000000", "#00000000", "#FF000000")
        or "None" in str(color)
        or color.upper().startswith("#00") and color.upper() not in ("#00FFFF", "#00FF00", "#0000FF")
    ):
        color = "#FFFFFF"

    # Debug-Ausgabe nur für Nicht-Weiß
    if color != "#FFFFFF":
        print(f"Zelle {cell.coordinate}: pattern={fill.patternType}, "
              f"fg={getattr(fill.fgColor, 'rgb', None)}, "
              f"bg={getattr(fill.bgColor, 'rgb', None)}, → {color}")

    return color

def read_excel_with_colors(filename, sheet_name=None, start_row=2, end_row=28):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    data, colors = [], []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        row_data, row_colors = [], []
        for cell in row:
            row_data.append(cell.value)
            row_colors.append(get_cell_color(wb, cell))
        data.append(row_data)
        colors.append(row_colors)
    return data, colors

# --- Daten einlesen und vorverarbeiten ---
raw_data, colors = read_excel_with_colors(EXCEL_FILE)

processed_data = []
current_season = None  # zuletzt gesehene Staffel

for row in raw_data:
    new_row = []
    # falls in Spalte A (Index 0) ein Staffelname steht, merken
    first_cell = row[0]
    if first_cell:
        current_season = str(first_cell).strip()

    for col_idx, cell in enumerate(row):
        if col_idx == 0:
            # Spalte A: enthält evtl. Staffelnamen
            new_row.append((str(cell) if cell else "", None, None))
        else:
            # andere Spalten: Episode innerhalb dieser Staffel
            ep_num = cell if cell is not None else ""
            new_row.append((str(ep_num), current_season, ep_num))
    processed_data.append(new_row)


# --- HTML-Template ---
html_template = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Season Numbering Tabelle</title>
<style>
/* Tabelle Desktop */
table { border-collapse: collapse; width: 100%; }
td { border: 1px solid #999; padding: 2px; text-align: center; font-size: 11px; }

/* Spalte B nur so breit wie nötig */
td.col-b {
    width: 1%;           /* minimaler Breitenwert */
    white-space: nowrap;  /* verhindert Zeilenumbruch */
}

/* Mobile Layout */
@media (max-width: 600px) {
  table, thead, tbody, th, td, tr { display: block; }
  tr { margin-bottom: 8px; border: 1px solid #999; border-radius: 4px; padding: 4px; }
  td { display: flex; justify-content: space-between; padding: 4px; border: none; border-bottom: 1px solid #ddd; }
  td:last-child { border-bottom: none; }
  td::before { content: attr(data-label); font-weight: bold; }
}
</style>
<script>
document.addEventListener("DOMContentLoaded", function() {
    // Popup-Container einmalig hinzufügen
    const popup = document.createElement("div");
    popup.id = "popup";
    popup.style.display = "none";
    popup.style.position = "absolute";
    popup.style.background = "#fff";
    popup.style.border = "1px solid #999";
    popup.style.padding = "10px";
    popup.style.borderRadius = "8px";
    popup.style.boxShadow = "0 4px 8px rgba(0,0,0,0.2)";
    popup.style.maxWidth = "300px";
    popup.style.zIndex = "9999";
    popup.innerHTML = `
        <div id="popup-content"></div>
        <button id="popup-close" style="margin-top:6px;">Schließen</button>
    `;
    document.body.appendChild(popup);

    // Schließen-Button
    document.getElementById("popup-close").addEventListener("click", () => {
        popup.style.display = "none";
    });

    // Klick auf Episode-Zelle
    document.querySelectorAll(".episode-cell").forEach(cell => {
        cell.addEventListener("click", function(event) {
            const season = this.dataset.season;
            const ep = this.dataset.ep;

            if (!season || !ep) return;

            fetch(`/get_code/${season}/${ep}`)
                .then(response => {
                    if (!response.ok) throw new Error("Serverfehler");
                    return response.text();
                })
                .then(code => {
                    // Popup-Inhalt einfügen
                    document.getElementById("popup-content").innerHTML = code;
                    
                    // Popup an Klickposition setzen
                    const x = event.pageX + 10; // 10px rechts vom Cursor
                    const y = event.pageY + 10; // 10px unter dem Cursor
                    popup.style.left = `${x}px`;
                    popup.style.top = `${y}px`;

                    popup.style.display = "block";
                })
                .catch(err => {
                    document.getElementById("popup-content").innerText = "Fehler: " + err;
                    popup.style.left = `${event.pageX + 10}px`;
                    popup.style.top = `${event.pageY + 10}px`;
                    popup.style.display = "block";
                });
        });
    });

    // Popup schließen, wenn man außerhalb klickt
    document.addEventListener("click", function(e) {
        if (!popup.contains(e.target) && !e.target.classList.contains("episode-cell")) {
            popup.style.display = "none";
        }
    });
});
</script>



</head>
<body>
<h2>Season Numbering Übersicht</h2>
<table>
{% for r in range(data|length) %}
    {% set row_data = data[r] %}
    {% set row_colors = colors[r] %}
    <tr>
      {% for c in range(0,2) %}
        {% set cell_value, season, ep_in_season = row_data[c] %}
        <td class="{{ 'col-b' if c == 1 else '' }} episode-cell"
            style="background-color: {{ row_colors[c] }};"
            data-season="{{ season if season is not none else '' }}"
            data-ep="{{ ep_in_season if ep_in_season is not none else '' }}">
            {{ cell_value if cell_value is not none else "" }}
        </td>
      {% endfor %}

      {% set end_index = 18 if row_data|length > 18 else row_data|length %}
      {% for c in range(2, end_index) %}
        {% set cell_value, season, ep_in_season = row_data[c] %}
        <td style="background-color: {{ row_colors[c] }};"
            data-season="{{ season if season is not none else '' }}"
            data-ep="{{ ep_in_season if ep_in_season is not none else '' }}"
            class="episode-cell">
        {{ cell_value if cell_value is not none else "" }}
        </td>

      {% endfor %}
    </tr>

    {% for start in range(18, row_data|length, 16) %}
      <tr>
        <td></td><td></td>
        {% for c in range(start, start + 16) %}
          {% if c < row_data|length %}
            {% set cell_value, season, ep_in_season = row_data[c] %}
            <td style="background-color: {{ row_colors[c] }};"
                data-season="{{ season if season is not none else '' }}"
                data-ep="{{ ep_in_season if ep_in_season is not none else '' }}"
                class="episode-cell">
            {{ cell_value if cell_value is not none else "" }}
            </td>

          {% endif %}
        {% endfor %}
      </tr>
    {% endfor %}
{% endfor %}
</table>
</body>
</html>
"""

# --- Flask-Routen ---
@app.route("/")
def index():
    return render_template_string(html_template, data=processed_data, colors=colors)

import json

# JSON beim Start laden
with open("data/episodes/episodes_events.json", "r", encoding="utf-8") as f:
    EPISODE_EVENTS = json.load(f)

@app.route("/get_code/<int:season>/<int:ep>")
def get_code(season, ep):
    try:
        episode_code = revert_episode_code(season, ep)

        # passenden Eintrag holen
        value = EPISODE_EVENTS.get(episode_code)
        if value is None:
            return f"Kein Eintrag für {episode_code}", 404

        # Wenn der Value eine Liste ist → formatieren als Aufzählung
        if isinstance(value, list):
            formatted = "<ul>" + "".join(f"<li>{item}</li>" for item in value) + "</ul>"
        else:
            formatted = str(value)

        # korrekt kodierte HTML-Antwort zurückgeben
        return formatted, 200, {"Content-Type": "text/html; charset=utf-8"}

    except Exception as e:
        return f"Fehler: {str(e)}", 400



# --- App starten ---
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5005, debug=True)
